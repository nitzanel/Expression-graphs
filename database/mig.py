'''
This code uses the openpyxl package for playing around with excel using Python code
to convert complete excel workbook (all sheets) to an SQLite database
The code assumes that the first row of every sheet is the column name
Every sheet is stored in a separate table
The sheet name is assigned as the table name for every sheet


Code will also change the version.txt file 
'''

import sqlite3
import openpyxl
from openpyxl import load_workbook
import re
import os
names = [name for name in os.listdir('.') if name.split('.')[1] == 'xlsx']

# check to see whats alrdy in :D
con = sqlite3.connect('db.db')

cursor = con.cursor()
cursor.execute('SELECT name FROM sqlite_master WHERE type="table";')
database_tables = cursor.fetchall()
tmp = []
for name in database_tables:
    tmp.append(name[0])
database_tables = tmp
tmp = []
for name in names:
    if name.split('.')[0] not in database_tables:
        tmp.append(name)
names = tmp

for name in names:
    query = 'CREATE TABLE ' + name.split('.')[0] + '(' 
    columns = []
    wb = load_workbook(filename=name,read_only=True)
    sheets = wb.get_sheet_names()
    sheet = wb[sheets[0]]
    for row in sheet.get_squared_range(min_row=1,max_row=1,min_col=1,max_col=sheet.max_column):
        for cell in row:
            # the values will be saved as text, for simplicity. convert them later.
            if cell.value in ['ID','gene_name']:
                query += cell.value + ' TEXT' + ', '
            else:
                query += cell.value + ' TEXT' + ', '
            columns.append(cell.value)
    query = query[:-2]
    query += ');'
    print query
    con.execute(query)

    tup = []
    for i, row in enumerate(sheet):
        tuprow = []
        if i == 0:
            continue
        for cell in row:
            tuprow.append(cell.value)
        tup.append(tuple(tuprow))
        # if there is more then one gene here, create a duplicate using the other names. will need revisit.
        if ',' in tup[-1][1]:
            last_tup = list(tup.pop())
            current_genes = last_tup[1].split(',')
            for gene in current_genes:
                values = [last_tup[0]] + [gene] + last_tup[2:]
                tup.append(tuple(values))
    insQuery1 = 'INSERT INTO ' + name.split('.')[0] + '('
    insQuery2 = ''
    for col in columns:
        insQuery1 += col + ', '
        insQuery2 += '?, '
    insQuery1 = insQuery1[:-2] + ') VALUES('
    insQuery2 = insQuery2[:-2] + ')'
    insQuery = insQuery1 + insQuery2
    print insQuery
    con.executemany(insQuery,tup)
    con.commit()
con.close()
# create version.txt file. change the version number everytime the database chagnes.

with open('version.txt','a') as f:
	pass
with open('version.txt','r+') as f:
	version = f.read()
	if version == '':
		version = 0.0
	else:
	  	version = float(version)
	f.seek(0)
	f.write(str(version +0.1))
	f.truncate()

