import openpyxl as pyxl
import numpy as np
import sqlite3 as lite
class Loader():
	def __init__(self):
		self.worksheets = {}
		self.cell_ranges = {}
		self.indexed = False
		self.cellIndexed = False
		self.db_name = 'database/db.db'
		self.tables_names = []
		self.is_open = False
		for item in self.get_tables_names():
			if 'exp' in item.split('_'):
				self.tables_names.append(item)
		
	# creates sql connection with the database.
	def setup(self):
		if not self.is_open:
			# Dangerous?
			# probably.
			self.conn = lite.connect(self.db_name, check_same_thread=False)
			print 'Opend',self.db_name
			self.is_open = True
	# close sql connection.
	def tear_down(self):
		if self.conn:
			self.conn.close()
			print 'closed',self.db_name
			self.is_open = False
	
	# autocomplete from sqlite database
	def autocomplete(self, gene_symbol):
		self.setup()
		# change after database update.
		query = ''.join(["SELECT gene_name from Female_Male_exp_levels_log2 WHERE gene_name LIKE '",gene_symbol,"%'"," OR gene_name LIKE '%_",gene_symbol,"%'"," LIMIT 20"])
		cursor = self.conn.execute(query)
		names = list(set(list(map(lambda x:x[0], cursor.fetchall()))))
		return names		
		

	# input: table name.
	# returns a list of columns names in a table.
	def get_columns_names(self,table_name):
		query = ' '.join(['SELECT * from',table_name])
		cursor = self.conn.execute(query)
		names = list(map(lambda x:x[0], cursor.description))
		return names

	# returns a list of names of the tables in the database
	def get_tables_names(self):
		self.setup()
		cursor = self.conn.cursor()
		cursor.execute('SELECT name FROM sqlite_master WHERE type="table";')
		tables_names = list(map(lambda x:x[0],cursor.fetchall()))
		self.tear_down()
		return tables_names

	# input:
	# condition: condition of the query, defaults to gene_name.
	# value: the value of the condition.
	# dataset: the table name.
	# cells: a cell type, if the value is ALL, will select all columns. default to ALL
	# output:
	# A query command according to input.
	def get_select_command(self,value,dataset,cells='ALL',condition='gene_name'):
		if cells == 'ALL':
			cells = '*'	
		else:
			cells = ', '.join(self.get_cells_names(cells,dataset))
		command = ' '.join(['SELECT',cells,'from',dataset,'where',condition,'=',''.join(['"',value,'"'])])
		return command

	# the function takes a gene_name, a list of datasets dirs, which can be set to 'ALL'
	# and an optional argument of cell type for cell specific graphs
	# returns a dictionary where the keys are the datasets and the values are lists of expression data 
	# for each row of the wgene in the database. 	
	def get_gene(self,gene_name,datasets='ALL',cells='ALL'):
		self.setup()
		data = {}
		if datasets == 'ALL':
			datasets = self.tables_names
		for dataset in datasets:
			values_list = self.get_gene_data(gene_name,dataset,cells)
			if cells == 'ALL':
				colms = self.get_columns_names(dataset)
			else:
				colms = self.get_cells_names(cells,dataset)
			data_tuples = {}
			for index,values in enumerate(values_list):
				key = '_'.join(['repeat',str(index+1)])
				data_tuples[key] = zip(colms,values)
			data[dataset] = data_tuples
		self.tear_down()
		return data

	# input:
	# cell_type: the type of cell, for example: 'GN','B1a'
	# table: the table name in the sql shema. get it from get_tables_names
	# returns a list of cells of the same type of the input cell.
	def get_cells_names(self,cell_type,table):
		cells = self.get_columns_names(table)
		cells_types = [cell_type]
		cells_names = []
		if cell_type.upper() == 'B1AB':
			cells_types.append('B1A')
		elif cell_type.upper() == 'CD19':
			cells_types.append('B')
		elif cell_type.upper() == 'T8':
			cells_types.append('CD8T')
		elif cell_type.upper() == 'T4':
			cells_types.append('CD4T')
		
		for cell in cells:
			for item in cells_types:
				if item.upper() in cell.upper().split('_'):
					cells_names.append(cell)
		return cells_names

	# input:
	# gene_name is the gene symbol
	# data_set is the table name in the sql database (the dataset)
	# cells is a list of cells to get values for (columns in the sql dataset). defaults to ALL.
	def get_gene_data(self,gene_name,data_set,cells='ALL'):	
		cursor = self.conn.execute(self.get_select_command(gene_name,data_set,cells))
		data = []
		for row in cursor:
			data.append(list(row))
		return data
 		

	def getCellsRanges(self):
		if self.cellIndexed is not True:
			self.getSets()
			for data_set in self.data_sets:
				if data_set == 'blank':
					continue
				
				if data_set not in self.cell_ranges:
					self.cell_ranges[data_set] = {}

				self.loadSheet(self.data_sets[data_set])
				head = self.loadHead(self.data_sets[data_set])
				last_cell = ''
				last_index = 0
				for index,info in enumerate(head):
					if index < 5:
						continue
					cell = info.split('_')[0]
					if cell != last_cell:
						if cell not in self.cell_ranges[data_set]:
							self.cell_ranges[data_set][cell] = {}
							self.cell_ranges[data_set][cell]['start'] = index+1
							if last_cell != '':
								self.cell_ranges[data_set][last_cell]['end'] = index
						last_cell = cell
					last_index = index
				self.cell_ranges[data_set][last_cell]['end'] = index+1
		return self.cell_ranges				




	# get the data sets list.
	def getSets(self):
		# create a list of the data sets.
		data_sets = {}
		self.blank_set = {}
		data_sets['blank'] = 'Blank.xlsx'
		data_sets['GenderExp'] = 'Female_Male_exp_levels_norm.xlsx'
		data_sets['Immgen'] = 'ImmGen_sex_exp_levels_norm.xlsx'
		self.data_sets = data_sets
		return data_sets

	# takes a gene name, returns the row index it's in.
	def findRowMatch(self,gene_name):
		if self.indexed is not True:
			# call loadNames with a worksheet that is already loaded.
			self.getSets()
			self.loadSheet(self.data_sets['blank'])
			self.loadNames(self.worksheets[self.data_sets['blank']]) 
		if gene_name in self.rows_dict:
			#print 'found gene '
			return self.rows_dict[gene_name]
		#return (''.join(['error ',gene_name, ' doesnt exist in the data set']))
		#print 'BAD GENE NAME'
		return -1

	# takes a list of workbook_names, and load the sheets into a dictionary for further use if they are not there.
	def loadSheet(self,workbook_name): 
		if workbook_name not in self.worksheets:
			# load the sheet into memory
			wb = pyxl.load_workbook(filename = workbook_name,read_only=True)
			self.worksheets[workbook_name] = wb.worksheets[0] # could use split to take the first part of the  _
	# takes a sheet, loads the gene column into a dictionary.
	def loadNames(self,sheet):
		self.rows_dict = {}
		for index, row in enumerate(sheet.get_squared_range(min_col=2,max_col=2,min_row=2,max_row=sheet.max_row)):
			for cell in row:
				self.rows_dict[cell.value] = index + 2
		self.indexed = True
		return self.rows_dict

	# loads the first row of the sheet into a list and returns it.
	def loadHead(self,sheet):
		info = []
		current_sheet = self.worksheets[sheet]
		for row in current_sheet.get_squared_range(min_col=1,max_col=current_sheet.max_column,min_row=1,max_row=1):
			for cell in row:
				info.append(cell.value)
		return info

	def loadRow(self,gene_name,sheet):
		row_index = self.findRowMatch(gene_name)
		row_data = []
		current_sheet = self.worksheets[sheet]
		for row in current_sheet.get_squared_range(min_col = 1,max_col = current_sheet.max_column,min_row=row_index,max_row=row_index):
			for cell in row:
				row_data.append(cell.value)
		return row_data

	def loadPartialRow(self,gene_name,cell_name,sheet,current_set):
		row_index = self.findRowMatch(gene_name)
		row_data = []
		current_sheet = self.worksheets[sheet]
		ranges = self.getCellsRanges()
		for row in current_sheet.get_squared_range(min_col=self.cell_ranges[current_set][cell_name]['start'],max_col = self.cell_ranges[current_set][cell_name]['end'],min_row=row_index,max_row=row_index):
			for cell in row:
				row_data.append(cell.value)
		return row_data

	# the function will return a dictionary of expression level of a specific gene in a specific cell in all the datasets.
	def loadCellSpecific(self,gene_name,cell_name):
		gene_cell_dict = {}
		sets_names = []
		if self.findRowMatch(gene_name) == -1:
			return -1

		for data_set in self.data_sets:
			if data_set == 'blank':
				continue
			self.loadSheet(self.data_sets[data_set])
		data = {}
		for data_set in self.data_sets:
			if data_set == 'blank':
				continue
			data[data_set] = self.loadPartialRow(gene_name,cell_name,self.data_sets[data_set],data_set)
		return data

	# takes a list of genes, loads them.
	# this will call the rest of the functions.
	# can improve reloading
	def loadGenes(self,genes_list,data_sets):
		data_sets_dict = {}
		for data_set in data_sets:
			self.loadSheet(data_set) 
			genes_dict = {}
			genes_dict['head'] = self.loadHead(data_set)
			for gene_name in genes_list:
				genes_dict[gene_name] = self.loadRow(gene_name,data_set)
			data_sets_dict[data_set] = genes_dict
		return data_sets_dict

